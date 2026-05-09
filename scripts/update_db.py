#!/usr/bin/env python3
"""Update MDM database from an XLSX file: add new records to JSONL and PostgreSQL with embeddings."""

import json
import os
import sys
import time

import openpyxl
import psycopg2
from psycopg2 import pool
import openai

# --- Config ---
XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else None
if not XLSX_PATH:
    print("Usage: python3 update_db.py <xlsx_file>")
    sys.exit(1)

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSONL_PATH = os.path.join(SKILL_DIR, 'data', 'mdm_nomenclature.jsonl')

# API key — from env var or .env file
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')
if not OPENAI_API_KEY:
    env_file = os.path.join(SKILL_DIR, '.env')
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            OPENAI_API_KEY = f.read().strip()
if not OPENAI_API_KEY:
    print("ERROR: OPENAI_API_KEY not set. Set env var or create .env file.", file=sys.stderr)
    sys.exit(1)

client = openai.OpenAI(api_key=OPENAI_API_KEY)

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_PORT = int(os.getenv('DB_PORT', 5432))
DB_NAME = os.getenv('DB_NAME', 'nomenclature_kb')
DB_USER = os.getenv('DB_USER', 'postgres')
DB_PASS = os.getenv('DB_PASS', '')

BATCH_SIZE = 50
EMBEDDING_MODEL = "text-embedding-3-small"

_db_pool = None

def get_db_pool():
    global _db_pool
    if _db_pool is None:
        _db_pool = pool.SimpleConnectionPool(
            1, 5,
            host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
            user=DB_USER, password=DB_PASS
        )
    return _db_pool

def get_embedding(text: str) -> list:
    resp = client.embeddings.create(model=EMBEDDING_MODEL, input=text)
    return resp.data[0].embedding

# --- Step 1: Load existing JSONL codes ---
print("=" * 60)
print("Step 1: Loading existing JSONL codes...")
existing_codes = set()
existing_records = []
if os.path.exists(JSONL_PATH):
    with open(JSONL_PATH, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                r = json.loads(line)
                existing_codes.add(r['code'])
                existing_records.append(r)
print(f"  Loaded {len(existing_codes)} existing codes")

# --- Step 2: Parse XLSX and find new records ---
print("\nStep 2: Parsing XLSX file...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb[wb.sheetnames[0]]

# Header: Краткое наименование, Полное наименование, Код, Класс, Классификационный код, Статус, Единица измерения, Артикул, ГТ
COL_MAP = {
    'name': 1,    # Краткое наименование
    'name_full': 2,  # Полное наименование
    'code': 3,    # Код
    'class': 4,   # Класс
    'class_code': 5,  # Классификационный код
    'status': 6,  # Статус
    'unit': 7,    # Единица измерения
    'article': 8, # Артикул
    'gt': 9,      # ГТ
}

new_records = []
for r in range(2, ws.max_row + 1):
    code = str(ws.cell(r, COL_MAP['code']).value).strip() if ws.cell(r, COL_MAP['code']).value else ''
    if not code:
        continue
    if code in existing_codes:
        continue

    name = str(ws.cell(r, COL_MAP['name']).value or '').strip()
    unit = str(ws.cell(r, COL_MAP['unit']).value or '').strip() or 'шт'

    record = {
        'name': name,
        'code': code,
        'class': str(ws.cell(r, COL_MAP['class']).value or '').strip(),
        'class_code': str(ws.cell(r, COL_MAP['class_code']).value or '').strip(),
        'status': str(ws.cell(r, COL_MAP['status']).value or '').strip() or 'Эталон',
        'unit': unit,
        'article': str(ws.cell(r, COL_MAP['article']).value or '').strip(),
        'gt': str(ws.cell(r, COL_MAP['gt']).value or '').strip(),
    }
    new_records.append(record)

print(f"  Found {len(new_records)} new records to add")

if not new_records:
    print("  Nothing to add — database is up to date!")
    sys.exit(0)

# --- Step 3: Append new records to JSONL ---
print("\nStep 3: Appending to JSONL...")
with open(JSONL_PATH, 'a', encoding='utf-8') as f:
    for rec in new_records:
        f.write(json.dumps(rec, ensure_ascii=False) + '\n')
print(f"  Appended {len(new_records)} records")
print(f"  Total JSONL: {len(existing_codes) + len(new_records)}")

# --- Step 4: Generate embeddings and insert into PostgreSQL ---
print(f"\nStep 4: Generating embeddings & inserting into PostgreSQL ({len(new_records)} records)...")

conn = get_db_pool().getconn()
cur = conn.cursor()

# Check which codes already have embeddings
codes_to_check = [r['code'] for r in new_records]
existing_emb_codes = set()
for i in range(0, len(codes_to_check), 500):
    batch = codes_to_check[i:i+500]
    cur.execute("SELECT code FROM mdm_reference WHERE code = ANY(%s)", (batch,))
    existing_emb_codes.update(row[0] for row in cur.fetchall())

records_to_embed = [r for r in new_records if r['code'] not in existing_emb_codes]
print(f"  Already in PG: {len(existing_emb_codes)}, need embeddings: {len(records_to_embed)}")

inserted = 0
for i in range(0, len(records_to_embed), BATCH_SIZE):
    batch = records_to_embed[i:i+BATCH_SIZE]
    
    # Generate embeddings
    texts = [r['name'] for r in batch]
    try:
        resp = client.embeddings.create(model=EMBEDDING_MODEL, input=texts)
        embeddings = [d.embedding for d in resp.data]
    except Exception as e:
        print(f"  Embedding error at batch {i}-{i+len(batch)}: {e}")
        time.sleep(2)
        continue

    # Insert
    for j, rec in enumerate(batch):
        try:
            cur.execute("""
                INSERT INTO mdm_reference (code, name, class, embedding, attributes)
                VALUES (%s, %s, %s, %s::vector, %s::jsonb)
                ON CONFLICT (code) DO UPDATE SET
                    name = EXCLUDED.name,
                    class = EXCLUDED.class,
                    embedding = EXCLUDED.embedding,
                    attributes = EXCLUDED.attributes
            """, (
                rec['code'],
                rec['name'],
                rec['class'],
                embeddings[j],
                json.dumps({
                    'status': rec['status'],
                    'unit': rec['unit'],
                    'class_code': rec['class_code'],
                    'article': rec['article'],
                    'gt': rec['gt'],
                }, ensure_ascii=False)
            ))
            inserted += 1
        except Exception as e:
            print(f"  Insert error for {rec['code']}: {e}")

    conn.commit()
    print(f"  Batch {i+1}-{min(i+BATCH_SIZE, len(records_to_embed))} / {len(records_to_embed)} — inserted: {inserted}")

print(f"  Total inserted/updated in PG: {inserted}")
cur.close()
get_db_pool().putconn(conn)

# --- Summary ---
print("\n" + "=" * 60)
print("UPDATE COMPLETE")
print(f"  New records in JSONL:  {len(new_records)}")
print(f"  New in PostgreSQL:     {inserted}")
print(f"  Total JSONL records:   {len(existing_codes) + len(new_records)}")
